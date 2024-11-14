import os
import fitz  # PyMuPDF for PDF page count
from docx import Document  # python-docx for Word
from openpyxl import load_workbook  # openpyxl for Excel
from pptx import Presentation  # python-pptx for PowerPoint

# Set limits
MAX_FILE_SIZE_MB = 5  # Maximum file size in MB
MAX_PAGE_LIMIT = 45   # Maximum content-based page limit

# Define functions to estimate page count for each file type

def get_pdf_page_count(file_path):
    """Estimate page count for a PDF file."""
    with fitz.open(file_path) as pdf:
        return pdf.page_count

def get_docx_page_count(file_path):
    """Estimate page count for a Word (DOCX) file."""
    doc = Document(file_path)
    word_count = sum(len(paragraph.text.split()) for paragraph in doc.paragraphs)
    # Estimate pages assuming ~300 words per page
    return word_count // 300 + (word_count % 300 > 0)

def get_pptx_page_count(file_path):
    """Estimate page count for a PowerPoint file (count slides as pages)."""
    ppt = Presentation(file_path)
    return len(ppt.slides)

def get_xlsx_page_count(file_path):
    """Estimate page count for an Excel file (count sheets as pages)."""
    workbook = load_workbook(file_path, data_only=True)
    return len(workbook.sheetnames)

def get_txt_page_count(file_path):
    """Estimate page count for a text file based on word count."""
    with open(file_path, 'r') as file:
        text = file.read()
    word_count = len(text.split())
    # Estimate pages assuming ~300 words per page
    return word_count // 300 + (word_count % 300 > 0)

# Determine file type and estimate page count

def estimate_page_count(file_path, file_type):
    """Estimate page count based on file type."""
    if file_type == 'pdf':
        return get_pdf_page_count(file_path)
    elif file_type == 'docx':
        return get_docx_page_count(file_path)
    elif file_type == 'pptx':
        return get_pptx_page_count(file_path)
    elif file_type == 'xlsx':
        return get_xlsx_page_count(file_path)
    elif file_type == 'txt':
        return get_txt_page_count(file_path)
    else:
        raise ValueError("Unsupported file type")

# Check file size

def check_file_size(file_path):
    """Check if the file size is within the allowed limit."""
    file_size_mb = os.path.getsize(file_path) / (1024 * 1024)  # Convert to MB
    return file_size_mb <= MAX_FILE_SIZE_MB

# Main validation function

def validate_file_upload(file_path, file_type):
    """
    Validate if a file meets both file size and content volume requirements.
    Raises an error if validation fails.
    """
    # Step 1: File size check
    if not check_file_size(file_path):
        raise ValueError(f"File exceeds the maximum allowed size of {MAX_FILE_SIZE_MB} MB.")
    
    # Step 2: Content-based volume check
    page_count = estimate_page_count(file_path, file_type)
    if page_count > MAX_PAGE_LIMIT:
        raise ValueError(f"File exceeds the maximum allowed {MAX_PAGE_LIMIT} pages (found {page_count} pages).")
    
    return True  # File is valid for upload

# Example usage
file_path = '/path/to/uploaded/file.pdf'  # Path to the file being uploaded
file_type = os.path.splitext(file_path)[-1].lower().replace('.', '')  # Extract file extension

try:
    if validate_file_upload(file_path, file_type):
        print("File is within the allowed limits and can be uploaded.")
except ValueError as e:
    print(str(e))
